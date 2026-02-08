"""Convert COMSCC Classing Sheets (Excel) to JSON."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Vehicles sheet layout
VEHICLES_ROW_START = 13
VEHICLES_ROW_END = 501
COL_MAKE = 1
COL_MODEL = 2
COL_START_YEAR = 3
COL_END_YEAR = 4
COL_SHOWROOM_BASE_WEIGHT = 5
COL_FACTORY_HP = 6
COL_FACTORY_TQ = 7
COL_SUSP_INDEX = 11

# Points sheet layout (Engine, Drivetrain, etc.)
COL_POINTS = 1
COL_DESCRIPTION = 2
POINTS_SHEET_ROW_START = 9
SHEET_ROW_RANGES: dict[str, tuple[int, int]] = {
    "Engine": (9, 58),
    "Drivetrain": (9, 30),
    "Suspension": (9, 38),
    "Brakes": (9, 19),
    "Exterior": (9, 29),
    "Tires": (9, 77),
}
SKIP_DESCRIPTION = "Dyno"


def _ascii_only(value: str) -> str:
    """Return string with non-ASCII characters removed."""
    return value.encode("ascii", "ignore").decode()


def process_vehicles(sheet: Worksheet) -> list[dict[str, Any]]:
    """Extract vehicle rows from the Vehicles sheet into a list of dicts."""
    entries = []
    for row_num, row in enumerate(
        sheet.iter_rows(
            min_row=VEHICLES_ROW_START,
            max_row=VEHICLES_ROW_END,
            values_only=True,
        ),
        start=VEHICLES_ROW_START,
    ):
        entries.append({
            "id": row_num,
            "make": row[COL_MAKE],
            "model": row[COL_MODEL],
            "start_year": row[COL_START_YEAR],
            "end_year": row[COL_END_YEAR],
            "showroom_weight": row[COL_SHOWROOM_BASE_WEIGHT],
            "factory_hp": row[COL_FACTORY_HP],
            "factory_tq": row[COL_FACTORY_TQ],
            "susp_index": row[COL_SUSP_INDEX],
        })
    return entries


def process_points_sheet(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
) -> list[dict[str, Any]]:
    """Extract points/description rows from a sheet into a list of dicts."""
    entries = []
    for row_num, row in enumerate(
        sheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            values_only=True,
        ),
        start=start_row,
    ):
        if row[COL_DESCRIPTION] == SKIP_DESCRIPTION:
            continue
        entries.append({
            "id": row_num,
            "description": _ascii_only(str(row[COL_DESCRIPTION])),
            "points": row[COL_POINTS],
        })
    return entries


def _write_json(data: list[dict[str, Any]], path: Path) -> None:
    """Write a list of dicts to a JSON file with indentation."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=4), encoding="utf-8")


def main() -> None:
    """CLI entry: convert workbook to JSON files."""
    parser = argparse.ArgumentParser(
        prog="sheetconverter",
        description="Converts COMSCC Classing Sheets to JSON",
    )
    parser.add_argument("-i", "--input", required=True, help="Input Excel file")
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=Path("./jsonfiles"),
        help="Output directory for JSON files (default: ./jsonfiles)",
    )
    args = parser.parse_args()

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(args.input, read_only=True, data_only=True)

    # Vehicles (different structure)
    vehicles_data = process_vehicles(wb["Vehicles"])
    _write_json(vehicles_data, output_dir / "vehicles.json")

    # Points-based sheets (same structure, different ranges)
    for sheet_name, (start_row, end_row) in SHEET_ROW_RANGES.items():
        data = process_points_sheet(wb[sheet_name], start_row, end_row)
        filename = sheet_name.lower() + ".json"
        _write_json(data, output_dir / filename)


if __name__ == "__main__":
    main()
